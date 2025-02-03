import openpyxl
import numpy as np


def parse_xlsx_exp(file_path, sheet_name):
    # Load the workbook and select the specified sheet
    wb = openpyxl.load_workbook(file_path)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' not found in the workbook.")

    sheet = wb[sheet_name]

    # List to hold the first row (headers)
    headers = []

    # List to hold parsed rows (dictionaries)
    parsed_rows = []

    # Iterate through the rows in the sheet
    for i, row in enumerate(sheet.iter_rows(values_only=True)):
        # The very first row is treated as headers
        if i == 0:
            meta = list(row)
            while meta and meta[-1] is None:
                meta.pop()
        elif i == 1:
            headers = list(row)
            while headers and headers[-1] is None:
                headers.pop()
        elif i > 1:
            # Create a dictionary for each row using the headers
            row_dict = {
                headers[j]: row[j] for j in range(len(row)) if row[j] is not None
            }
            parsed_rows.append(row_dict)

    parsed_rows = [row for row in parsed_rows if row]

    return meta, parsed_rows


def parse_xlsx_mice(file_path, sheet_name="mice info"):
    # Load the workbook and select the specified sheet
    wb = openpyxl.load_workbook(file_path)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' not found in the workbook.")

    sheet = wb[sheet_name]

    # List to hold the first row (headers)
    headers = []

    # List to hold parsed rows (dictionaries)
    parsed_rows = []

    # Iterate through the rows in the sheet
    for i, row in enumerate(sheet.iter_rows(values_only=True)):
        # The very first row is treated as headers
        if i == 0:
            headers = list(row)
            while headers and headers[-1] is None:
                headers.pop()
        else:
            # Create a dictionary for each row using the headers
            row_dict = {
                headers[j]: row[j] for j in range(len(row)) if row[j] is not None
            }
            parsed_rows.append(row_dict)

    parsed_rows = [row for row in parsed_rows if row]

    return parsed_rows


def save_row_as_npy(exp, mice_info, i, mouse_name=None):

    try:
        if not mouse_name:
            mouse_name = int(mice_info["mouse_ID"])

        # Define the dictionary template with keys
        # template = ['mouse_name', 'start_date', 'day', 'mouse_id', 'dob', 'sex', 'last_exp', 'strain', 'doc', 'doe', 'experimenter_name', 'attempt', 'weight_percentage', 'session_notes', 'rig_id', 'task_name', 'license', 'anesthesia_name', 'body_condition', 'general_assay', 'housing_assay', 'opto_name', 'opto_variant_name', 'opto_timing_name', 'opto_region_name']

        # headers default
        # headers = ['date', 'weight', 'drinko', 'task', 'time', 'trials', 'rewards', 'reward rate', 'reward size/ul', 'task water', 'remain water', 'driko after adding', 'actual water intake after task', 'total intake', 'experimenter_name', 'body_condition', 'general_assay', 'housing_assay', 'anesthesia_name']

        date = exp["date"].strftime("%Y-%m-%d")
        sex = (
            "F"
            if mice_info["sex"] == "Female"
            else "M"
            if mice_info["sex"] == "Male"
            else mice_info["sex"]
        )
        weight_percentage = round(
            (float(exp["weight"]) - float(mice_info["baseline_weight"]))
            / float(mice_info["baseline_weight"])
            * 100,
            2,
        )

        output = {
            "mouse_name": mouse_name,
            "start_date": None,
            "day": None,
            "mouse_id": int(mice_info["mouse_ID"]),
            "dob": mice_info["DOB"].strftime("%Y-%m-%d"),
            "sex": sex,
            "last_exp": None,
            "strain": mice_info["strain"],
            "doc": date,
            "doe": date,
            "experimenter_name": exp["experimenter_name"],
            "attempt": 1,
            "weight_percentage": weight_percentage,
            "session_notes": None,
            "rig_id": None,
            "task_name": "AR_visual_discrimination",
            "license": None,
            "anesthesia_name": exp["anesthesia_name"],
            "body_condition": exp["body_condition"],
            "general_assay": exp["general_assay"],
            "housing_assay": exp["housing_assay"],
            "opto_name": "none",
            "opto_variant_name": "none",
            "opto_timing_name": "none",
            "opto_region_name": "none",
        }

        np.save(
            f"{output['mouse_name']}_{output['doe']}_{output['attempt']}.npy", output
        )

    except Exception as e:
        print(f"An error occurred: {e} for {mouse_name} row {i}")


# Example usage
file_path = "mice_weights.xlsx"
mice_info = parse_xlsx_mice(file_path)

for mouse in mice_info:
    meta, exp = parse_xlsx_exp(file_path, str(int(mouse["mouse_ID"])))
    i = 3
    for session in exp:
        save_row_as_npy(session, mouse, i)
        i += 1
